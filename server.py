from __future__ import annotations

import base64
import hashlib
import hmac
import io
import json
import os
import secrets
import sqlite3
import sys
from datetime import datetime, timezone
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from email import policy
from email.parser import BytesParser

from openpyxl import Workbook, load_workbook


APP_DIR = Path(__file__).resolve().parent
STATIC_DIR = APP_DIR / "static"
DATA_DIR = APP_DIR / "data"
DB_PATH = DATA_DIR / "app.db"
SESSION_COOKIE = "carton_session"
PASSWORD_ITERATIONS = 260_000

CATEGORIES = {
    "Direct Outer": 0.90,
    "Thermocol Box": 0.90,
    "Temperature Thermocol Box": 0.70,
    "Validated Cold Chain Box": 0.90,
}

CATEGORY_ALIASES = {
    "direct outer": "Direct Outer",
    "thermocol box": "Thermocol Box",
    "temperature thermocol box": "Temperature Thermocol Box",
    "validated cold chain box": "Validated Cold Chain Box",
    "cold chain box": "Validated Cold Chain Box",
}

INITIAL_USERS = [
    ("admin", "admin"),
    ("priyanka", "user"),
    ("vandana", "user"),
    ("kalpna", "user"),
    ("poonam", "user"),
    ("meghna", "user"),
    ("sumit", "user"),
    ("jisha", "user"),
    ("larissa", "user"),
]

DEFAULT_PASSWORD = "ChangeMe123!"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt, PASSWORD_ITERATIONS
    )
    return "pbkdf2_sha256${}${}${}".format(
        PASSWORD_ITERATIONS,
        base64.b64encode(salt).decode("ascii"),
        base64.b64encode(digest).decode("ascii"),
    )


def verify_password(password: str, stored: str) -> bool:
    try:
        method, iterations, salt_b64, digest_b64 = stored.split("$", 3)
        if method != "pbkdf2_sha256":
            return False
        salt = base64.b64decode(salt_b64)
        expected = base64.b64decode(digest_b64)
        actual = hashlib.pbkdf2_hmac(
            "sha256", password.encode("utf-8"), salt, int(iterations)
        )
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                role TEXT NOT NULL CHECK(role IN ('admin', 'user')),
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS cartons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                category TEXT NOT NULL,
                length REAL NOT NULL,
                breadth REAL NOT NULL,
                height REAL NOT NULL,
                volume REAL NOT NULL,
                tare_weight REAL NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS calculations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                reference TEXT NOT NULL,
                category TEXT NOT NULL,
                result_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            """
        )
        for username, role in INITIAL_USERS:
            exists = conn.execute(
                "SELECT id FROM users WHERE username = ?", (username,)
            ).fetchone()
            if exists:
                continue
            conn.execute(
                """
                INSERT INTO users (username, role, password_hash, created_at)
                VALUES (?, ?, ?, ?)
                """,
                (username, role, hash_password(DEFAULT_PASSWORD), now_iso()),
            )


def json_response(handler, payload, status=HTTPStatus.OK):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def text_response(handler, body: str, status=HTTPStatus.OK):
    encoded = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "text/plain; charset=utf-8")
    handler.send_header("Content-Length", str(len(encoded)))
    handler.end_headers()
    handler.wfile.write(encoded)


def read_json(handler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length) if length else b"{}"
    if not raw:
        return {}
    try:
        return json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError:
        raise ValueError("Invalid JSON body.")


def parse_upload(handler):
    """Read the uploaded Excel file without the removed Python 3.13 cgi module."""
    content_type = handler.headers.get("Content-Type", "")
    if "multipart/form-data" not in content_type.lower():
        raise ValueError("Please upload an Excel file.")

    try:
        length = int(handler.headers.get("Content-Length", "0"))
    except ValueError:
        length = 0
    if length <= 0:
        raise ValueError("Please upload an Excel file.")

    body = handler.rfile.read(length)
    message = BytesParser(policy=policy.default).parsebytes(
        (f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n").encode("utf-8")
        + body
    )
    if not message.is_multipart():
        raise ValueError("Please upload an Excel file.")

    for part in message.iter_parts():
        field_name = part.get_param("name", header="content-disposition")
        if field_name == "file":
            payload = part.get_payload(decode=True)
            if payload:
                return payload
    raise ValueError("Please upload an Excel file.")


def current_user(handler):
    cookie_header = handler.headers.get("Cookie", "")
    cookies = SimpleCookie(cookie_header)
    morsel = cookies.get(SESSION_COOKIE)
    if not morsel:
        return None
    token = morsel.value
    with db() as conn:
        row = conn.execute(
            """
            SELECT users.id, users.username, users.role
            FROM sessions
            JOIN users ON users.id = sessions.user_id
            WHERE sessions.token = ?
            """,
            (token,),
        ).fetchone()
        return dict(row) if row else None


def require_user(handler):
    user = current_user(handler)
    if not user:
        json_response(handler, {"error": "Please log in."}, HTTPStatus.UNAUTHORIZED)
        return None
    return user


def require_admin(handler):
    user = require_user(handler)
    if not user:
        return None
    if user["role"] != "admin":
        json_response(handler, {"error": "Admin access required."}, HTTPStatus.FORBIDDEN)
        return None
    return user


def normalize_category(value) -> str:
    category = str(value or "").strip()
    key = " ".join(category.lower().split())
    if key in CATEGORY_ALIASES:
        return CATEGORY_ALIASES[key]
    raise ValueError(
        "Invalid category '{}'. Use: {}.".format(category, ", ".join(CATEGORIES))
    )


def number(value, field: str, allow_blank=False):
    if value is None or str(value).strip() == "":
        if allow_blank:
            return None
        raise ValueError(f"{field} is required.")
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field} must be a number.")
    if parsed <= 0:
        raise ValueError(f"{field} must be greater than zero.")
    return parsed


def get_header_map(sheet, row_num=1):
    headers = {}
    for idx, cell in enumerate(
        next(sheet.iter_rows(min_row=row_num, max_row=row_num)), start=1
    ):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = idx
    return headers


def required_column(headers, *names):
    for name in names:
        key = name.strip().lower()
        if key in headers:
            return headers[key]
    raise ValueError(f"Missing required column: {names[0]}")


def find_carton_table(workbook):
    required_headers = {
        "code",
        "category",
        "length",
        "breadth",
        "height",
        "volume",
    }
    tare_headers = {"tare weight", "tare weight (g)"}

    for sheet in workbook.worksheets:
        for row_num in range(1, min(sheet.max_row, 100) + 1):
            headers = get_header_map(sheet, row_num)
            if required_headers.issubset(headers) and tare_headers.intersection(headers):
                return sheet, row_num, headers

    raise ValueError(
        "Could not find the carton master headers. Include Code, Category, Length, "
        "Breadth, Height, Volume, and Tare Weight in one row."
    )


def parse_carton_workbook(raw: bytes):
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    sheet, header_row, headers = find_carton_table(wb)
    columns = {
        "code": required_column(headers, "Code"),
        "category": required_column(headers, "Category"),
        "length": required_column(headers, "Length"),
        "breadth": required_column(headers, "Breadth"),
        "height": required_column(headers, "Height"),
        "volume": required_column(headers, "Volume"),
        "tare_weight": required_column(headers, "Tare Weight", "Tare Weight (g)"),
    }
    cartons = []
    seen_codes = set()
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1), start=header_row + 1
    ):
        values = {name: row[col - 1].value for name, col in columns.items()}
        if all(value in (None, "") for value in values.values()):
            continue
        code = str(values["code"] or "").strip()
        if not code:
            raise ValueError(f"Row {row_num}: Code is required.")
        if code.lower() in seen_codes:
            raise ValueError(f"Row {row_num}: Duplicate carton code '{code}'.")
        seen_codes.add(code.lower())
        cartons.append(
            {
                "code": code,
                "category": normalize_category(values["category"]),
                "length": number(values["length"], f"Row {row_num} Length"),
                "breadth": number(values["breadth"], f"Row {row_num} Breadth"),
                "height": number(values["height"], f"Row {row_num} Height"),
                "volume": number(values["volume"], f"Row {row_num} Volume"),
                "tare_weight": number(
                    values["tare_weight"], f"Row {row_num} Tare Weight"
                ),
            }
        )
    if not cartons:
        raise ValueError("The carton master file has no carton rows.")
    return cartons


def parse_items_workbook(raw: bytes):
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    sheet = wb.worksheets[0]
    headers = get_header_map(sheet)
    columns = {
        "name": required_column(headers, "Item Name"),
        "length": required_column(headers, "Length"),
        "breadth": required_column(headers, "Breadth"),
        "height": required_column(headers, "Height"),
        "quantity": required_column(headers, "Quantity"),
        "weight_per_unit": required_column(
            headers, "Weight per Unit (g)", "Weight Per Unit (g)"
        ),
    }
    items = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {name: row[col - 1].value for name, col in columns.items()}
        if all(value in (None, "") for value in values.values()):
            continue
        item = clean_item(values, row_num)
        items.append(item)
    if not items:
        raise ValueError("The item file has no item rows.")
    return items


def clean_item(raw: dict, row_num=None):
    prefix = f"Row {row_num}: " if row_num else ""
    name = str(raw.get("name") or raw.get("itemName") or "").strip()
    if not name:
        raise ValueError(prefix + "Item Name is required.")
    quantity = number(raw.get("quantity"), prefix + "Quantity")
    if int(quantity) != quantity:
        raise ValueError(prefix + "Quantity must be a whole number.")
    weight = number(raw.get("weight_per_unit"), prefix + "Weight per Unit (g)", True)
    return {
        "name": name,
        "length": number(raw.get("length"), prefix + "Length"),
        "breadth": number(raw.get("breadth"), prefix + "Breadth"),
        "height": number(raw.get("height"), prefix + "Height"),
        "quantity": int(quantity),
        "weight_per_unit": weight,
    }


def calculate(reference: str, category: str, raw_items: list[dict]):
    reference = str(reference or "").strip()
    if not reference:
        raise ValueError("Shipment reference is required.")
    category = normalize_category(category)
    items = [clean_item(item) for item in raw_items]
    if not items:
        raise ValueError("Please add at least one item.")

    with db() as conn:
        carton_rows = conn.execute(
            "SELECT * FROM cartons WHERE category = ? ORDER BY volume DESC", (category,)
        ).fetchall()
    if not carton_rows:
        raise ValueError(f"No cartons found for category '{category}'.")

    cartons = []
    efficiency = CATEGORIES[category]
    for row in carton_rows:
        carton = dict(row)
        carton["usable_capacity"] = carton["volume"] * efficiency
        cartons.append(carton)
    cartons.sort(key=lambda item: item["usable_capacity"], reverse=True)

    item_details = []
    total_item_volume = 0.0
    total_item_weight_g = 0.0
    all_weights_present = True
    for item in items:
        item_volume = item["length"] * item["breadth"] * item["height"] * item["quantity"]
        item_weight_g = None
        if item["weight_per_unit"] is None:
            all_weights_present = False
        else:
            item_weight_g = item["weight_per_unit"] * item["quantity"]
            total_item_weight_g += item_weight_g
        total_item_volume += item_volume
        item_details.append({**item, "volume": item_volume, "total_weight_g": item_weight_g})

    largest = cartons[0]
    selected = []
    if total_item_volume <= largest["usable_capacity"]:
        fitting = [
            carton for carton in cartons if carton["usable_capacity"] >= total_item_volume
        ]
        chosen = min(fitting, key=lambda carton: carton["usable_capacity"])
        selected.append({**chosen, "quantity": 1})
    else:
        full_count = int(total_item_volume // largest["usable_capacity"])
        remainder = total_item_volume - (full_count * largest["usable_capacity"])
        if full_count:
            selected.append({**largest, "quantity": full_count})
        if remainder > 0.000001:
            fitting = [
                carton for carton in cartons if carton["usable_capacity"] >= remainder
            ]
            chosen = min(fitting, key=lambda carton: carton["usable_capacity"]) if fitting else largest
            existing = next(
                (line for line in selected if line["code"] == chosen["code"]), None
            )
            if existing:
                existing["quantity"] += 1
            else:
                selected.append({**chosen, "quantity": 1})

    carton_summary = []
    total_cartons = 0
    total_carton_volume = 0.0
    total_tare_weight_g = 0.0
    for carton in selected:
        qty = int(carton["quantity"])
        total_cartons += qty
        total_carton_volume += carton["volume"] * qty
        total_tare_weight_g += carton["tare_weight"] * qty
        carton_summary.append(
            {
                "code": carton["code"],
                "category": carton["category"],
                "quantity": qty,
                "volume": carton["volume"],
                "usable_capacity": carton["usable_capacity"],
                "tare_weight": carton["tare_weight"],
            }
        )

    gross_weight_kg = None
    if all_weights_present:
        gross_weight_kg = (total_item_weight_g + total_tare_weight_g) / 1000

    return {
        "reference": reference,
        "category": category,
        "usablePercentage": int(efficiency * 100),
        "items": item_details,
        "cartons": carton_summary,
        "totals": {
            "itemVolume": total_item_volume,
            "cartons": total_cartons,
            "cartonVolume": total_carton_volume,
            "itemWeightKg": total_item_weight_g / 1000 if all_weights_present else None,
            "tareWeightKg": total_tare_weight_g / 1000,
            "grossWeightKg": gross_weight_kg,
            "volumeWeightKg": total_carton_volume / 6000,
            "weightsComplete": all_weights_present,
        },
    }


def workbook_bytes(wb: Workbook) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def send_xlsx(handler, filename: str, body: bytes):
    handler.send_response(HTTPStatus.OK)
    handler.send_header(
        "Content-Type",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    handler.send_header("Content-Disposition", f'attachment; filename="{filename}"')
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def build_carton_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Carton Master"
    ws.append(["Code", "Category", "Length", "Breadth", "Height", "Volume", "Tare Weight"])
    ws.append(["A1", "Direct Outer", 40, 30, 25, 30000, 450])
    ws.append(["T1", "Temperature Thermocol Box", 35, 25, 20, 17500, 600])
    ws.append(["V1", "Validated Cold Chain Box", 45, 35, 30, 47250, 900])
    return workbook_bytes(wb)


def build_item_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    ws.append(["Item Name", "Length", "Breadth", "Height", "Quantity", "Weight per Unit (g)"])
    ws.append(["Medicine A", 10, 5, 4, 25, 50])
    ws.append(["Medicine B", 12, 8, 6, 10, ""])
    return workbook_bytes(wb)


def build_result_workbook(result: dict, username: str, created_at: str):
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    totals = result["totals"]
    summary.append(["Shipment Reference", result["reference"]])
    summary.append(["Created By", username])
    summary.append(["Created At", created_at])
    summary.append(["Category", result["category"]])
    summary.append(["Usable Capacity", f"{result['usablePercentage']}%"])
    summary.append(["Total Item Volume", round(totals["itemVolume"], 2)])
    summary.append(["Total Boxes", totals["cartons"]])
    summary.append(["Total Carton Volume", round(totals["cartonVolume"], 2)])
    summary.append(["Total Volume Weight (kg)", round(totals["volumeWeightKg"], 2)])
    if totals["grossWeightKg"] is None:
        summary.append(["Total Gross Weight (kg)", "Not calculated - item weight missing"])
    else:
        summary.append(["Total Gross Weight (kg)", round(totals["grossWeightKg"], 2)])
        summary.append(["Total Item Weight (kg)", round(totals["itemWeightKg"], 2)])
    summary.append(["Total Carton Tare Weight (kg)", round(totals["tareWeightKg"], 2)])

    carton_sheet = wb.create_sheet("Cartons")
    carton_sheet.append(
        ["Code", "Category", "Quantity", "Carton Volume", "Usable Capacity", "Tare Weight"]
    )
    for carton in result["cartons"]:
        carton_sheet.append(
            [
                carton["code"],
                carton["category"],
                carton["quantity"],
                carton["volume"],
                carton["usable_capacity"],
                carton["tare_weight"],
            ]
        )

    item_sheet = wb.create_sheet("Items")
    item_sheet.append(
        [
            "Item Name",
            "Length",
            "Breadth",
            "Height",
            "Quantity",
            "Weight per Unit (g)",
            "Item Volume",
            "Total Weight (g)",
        ]
    )
    for item in result["items"]:
        item_sheet.append(
            [
                item["name"],
                item["length"],
                item["breadth"],
                item["height"],
                item["quantity"],
                item["weight_per_unit"],
                item["volume"],
                item["total_weight_g"],
            ]
        )
    return workbook_bytes(wb)


class AppHandler(SimpleHTTPRequestHandler):
    server_version = "CartonCalculator/1.0"

    def translate_path(self, path):
        parsed = urlparse(path)
        clean_path = parsed.path
        if clean_path == "/":
            return str(STATIC_DIR / "index.html")
        return str(STATIC_DIR / clean_path.lstrip("/"))

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/api/me":
                user = current_user(self)
                json_response(self, {"user": user, "categories": CATEGORIES})
            elif path == "/api/users":
                self.handle_users()
            elif path == "/api/cartons":
                self.handle_cartons()
            elif path == "/api/calculations":
                self.handle_calculations()
            elif path.startswith("/api/calculations/") and path.endswith("/export"):
                self.handle_export(path)
            elif path == "/api/templates/cartons":
                send_xlsx(self, "carton-master-template.xlsx", build_carton_template())
            elif path == "/api/templates/items":
                send_xlsx(self, "item-upload-template.xlsx", build_item_template())
            elif path.startswith("/api/"):
                json_response(self, {"error": "Not found."}, HTTPStatus.NOT_FOUND)
            else:
                super().do_GET()
        except Exception as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            import re as _re
            m = _re.match(r"^/api/calculations/(\d+)$", path)
            if m:
                self.handle_delete_calculation(int(m.group(1)))
            else:
                json_response(self, {"error": "Not found."}, HTTPStatus.NOT_FOUND)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_delete_calculation(self, calc_id: int):
        user = require_user(self)
        if user is None:
            return
        with db() as conn:
            calc = self.get_calculation(conn, user, calc_id)
            if calc is None:
                json_response(self, {"error": "Not found."}, HTTPStatus.NOT_FOUND)
                return
            conn.execute("DELETE FROM calculations WHERE id = ?", (calc_id,))
        json_response(self, {"ok": True})

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/api/login":
                self.handle_login()
            elif path == "/api/logout":
                self.handle_logout()
            elif path == "/api/change-password":
                self.handle_change_password()
            elif path == "/api/users/reset-password":
                self.handle_reset_password()
            elif path == "/api/cartons/upload":
                self.handle_carton_upload()
            elif path == "/api/items/upload":
                self.handle_item_upload()
            elif path == "/api/calculate":
                self.handle_calculate()
            else:
                json_response(self, {"error": "Not found."}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_login(self):
        payload = read_json(self)
        username = str(payload.get("username", "")).strip().lower()
        password = str(payload.get("password", ""))
        with db() as conn:
            row = conn.execute(
                "SELECT id, username, role, password_hash FROM users WHERE username = ?",
                (username,),
            ).fetchone()
            if not row or not verify_password(password, row["password_hash"]):
                json_response(
                    self, {"error": "Invalid username or password."}, HTTPStatus.UNAUTHORIZED
                )
                return
            token = secrets.token_urlsafe(32)
            conn.execute(
                "INSERT INTO sessions (token, user_id, created_at) VALUES (?, ?, ?)",
                (token, row["id"], now_iso()),
            )
        user = {"id": row["id"], "username": row["username"], "role": row["role"]}
        body = json.dumps({"user": user, "categories": CATEGORIES}).encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE}={token}; HttpOnly; SameSite=Lax; Path=/",
        )
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def handle_logout(self):
        cookies = SimpleCookie(self.headers.get("Cookie", ""))
        morsel = cookies.get(SESSION_COOKIE)
        if morsel:
            with db() as conn:
                conn.execute("DELETE FROM sessions WHERE token = ?", (morsel.value,))
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE}=; HttpOnly; SameSite=Lax; Path=/; Max-Age=0",
        )
        self.send_header("Content-Type", "application/json")
        self.end_headers()
        self.wfile.write(b'{"ok": true}')

    def handle_change_password(self):
        user = require_user(self)
        if not user:
            return
        payload = read_json(self)
        current = str(payload.get("currentPassword", ""))
        new_password = str(payload.get("newPassword", ""))
        if len(new_password) < 8:
            raise ValueError("New password must be at least 8 characters.")
        with db() as conn:
            row = conn.execute(
                "SELECT password_hash FROM users WHERE id = ?", (user["id"],)
            ).fetchone()
            if not verify_password(current, row["password_hash"]):
                raise ValueError("Current password is incorrect.")
            conn.execute(
                "UPDATE users SET password_hash = ? WHERE id = ?",
                (hash_password(new_password), user["id"]),
            )
        json_response(self, {"ok": True})

    def handle_reset_password(self):
        admin = require_admin(self)
        if not admin:
            return
        payload = read_json(self)
        username = str(payload.get("username", "")).strip().lower()
        new_password = str(payload.get("newPassword", ""))
        if len(new_password) < 8:
            raise ValueError("New password must be at least 8 characters.")
        with db() as conn:
            cur = conn.execute(
                "UPDATE users SET password_hash = ? WHERE username = ?",
                (hash_password(new_password), username),
            )
            if cur.rowcount == 0:
                raise ValueError("User not found.")
        json_response(self, {"ok": True})

    def handle_users(self):
        admin = require_admin(self)
        if not admin:
            return
        with db() as conn:
            rows = conn.execute(
                "SELECT username, role, created_at FROM users ORDER BY role, username"
            ).fetchall()
        json_response(self, {"users": [dict(row) for row in rows]})

    def handle_carton_upload(self):
        admin = require_admin(self)
        if not admin:
            return
        raw = parse_upload(self)
        cartons = parse_carton_workbook(raw)
        with db() as conn:
            conn.execute("DELETE FROM cartons")
            conn.executemany(
                """
                INSERT INTO cartons
                    (code, category, length, breadth, height, volume, tare_weight, created_at)
                VALUES
                    (:code, :category, :length, :breadth, :height, :volume, :tare_weight, :created_at)
                """,
                [{**carton, "created_at": now_iso()} for carton in cartons],
            )
        json_response(self, {"ok": True, "count": len(cartons)})

    def handle_cartons(self):
        user = require_user(self)
        if not user:
            return
        with db() as conn:
            rows = conn.execute(
                "SELECT code, category, length, breadth, height, volume, tare_weight FROM cartons ORDER BY category, volume DESC"
            ).fetchall()
        json_response(self, {"cartons": [dict(row) for row in rows]})

    def handle_item_upload(self):
        user = require_user(self)
        if not user:
            return
        raw = parse_upload(self)
        json_response(self, {"items": parse_items_workbook(raw)})

    def handle_calculate(self):
        user = require_user(self)
        if not user:
            return
        payload = read_json(self)
        result = calculate(
            payload.get("reference", ""),
            payload.get("category", ""),
            payload.get("items", []),
        )
        saved_id = None
        if payload.get("save"):
            with db() as conn:
                cur = conn.execute(
                    """
                    INSERT INTO calculations (user_id, reference, category, result_json, created_at)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        user["id"],
                        result["reference"],
                        result["category"],
                        json.dumps(result),
                        now_iso(),
                    ),
                )
                saved_id = cur.lastrowid
        json_response(self, {"result": result, "savedId": saved_id})

    def handle_calculations(self):
        user = require_user(self)
        if not user:
            return
        params = parse_qs(urlparse(self.path).query)
        with db() as conn:
            if "id" in params:
                row = self.get_calculation(conn, user, int(params["id"][0]))
                json_response(self, {"calculation": self.calc_row(row)})
                return
            if user["role"] == "admin":
                rows = conn.execute(
                    """
                    SELECT calculations.*, users.username
                    FROM calculations
                    JOIN users ON users.id = calculations.user_id
                    ORDER BY calculations.created_at DESC
                    """
                ).fetchall()
            else:
                rows = conn.execute(
                    """
                    SELECT calculations.*, users.username
                    FROM calculations
                    JOIN users ON users.id = calculations.user_id
                    WHERE calculations.user_id = ?
                    ORDER BY calculations.created_at DESC
                    """,
                    (user["id"],),
                ).fetchall()
        json_response(self, {"calculations": [self.calc_row(row, False) for row in rows]})

    def handle_export(self, path: str):
        user = require_user(self)
        if not user:
            return
        parts = path.strip("/").split("/")
        calc_id = int(parts[2])
        with db() as conn:
            row = self.get_calculation(conn, user, calc_id)
        result = json.loads(row["result_json"])
        body = build_result_workbook(result, row["username"], row["created_at"])
        safe_ref = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in result["reference"])
        send_xlsx(self, f"{safe_ref or 'calculation'}-carton-result.xlsx", body)

    def get_calculation(self, conn, user, calc_id: int):
        if user["role"] == "admin":
            row = conn.execute(
                """
                SELECT calculations.*, users.username
                FROM calculations
                JOIN users ON users.id = calculations.user_id
                WHERE calculations.id = ?
                """,
                (calc_id,),
            ).fetchone()
        else:
            row = conn.execute(
                """
                SELECT calculations.*, users.username
                FROM calculations
                JOIN users ON users.id = calculations.user_id
                WHERE calculations.id = ? AND calculations.user_id = ?
                """,
                (calc_id, user["id"]),
            ).fetchone()
        if not row:
            raise ValueError("Calculation not found.")
        return row

    def calc_row(self, row, include_result=True):
        data = {
            "id": row["id"],
            "username": row["username"],
            "reference": row["reference"],
            "category": row["category"],
            "created_at": row["created_at"],
        }
        if include_result:
            data["result"] = json.loads(row["result_json"])
        return data


def main():
    init_db()
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer(("0.0.0.0", port), AppHandler)
    print(f"Carton Calculator running at http://127.0.0.1:{port}")
    server.serve_forever()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
    except Exception as exc:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        (DATA_DIR / "startup-error.log").write_text(repr(exc), encoding="utf-8")
        raise
